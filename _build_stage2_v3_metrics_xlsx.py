"""
Build an xlsx summary of ALL per-iteration metrics from results/
mfdro_stage2_v3/checkpoints, one sheet per benchmark, long format (one row
per method/seed/iteration). Only reads the worker-level output files
({METHOD}__{benchmark}__seed{N}.json) -- excludes SF-DRO's own internal
resumability checkpoints ({benchmark}__SF-DRO-rotate-MES__seed{N}.json /
Ackley_10D_MF__...) and MF-DRO's .mf.json duplicates, which are a separate
checkpoint scheme, not additional runs.
"""
import json
import glob
import os
import re

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

CKPT_DIR = "results/mfdro_stage2_v3/checkpoints"
OUT_PATH = "results/mfdro_stage2_v3/mfdro_stage2_v3_metrics_by_iteration.xlsx"

BENCHMARKS = ["Currin_2D", "Hartmann_6D", "Borehole_8D", "Ackley_10D"]
METHODS = ["MF-DRO", "SF-DRO", "Greedy-MES"]
SEEDS = [42, 43, 44, 45, 46]

# per-iteration array key -> unified column name
ARRAY_COLS = [
    ("hf_regret_curve", "regret"), ("regret_curve", "regret"),
    ("cost_curve", "cost"),
    ("cumulative_cost_curve", "cumulative_cost"),
    ("fidelity_trace", "fidelity_H1_L0"),
    ("y_t_trace", "y_t"),
    ("x_t_trace", "x_t"),
    ("rtg_target", "rtg_target"),
    ("btg_target", "btg_target"),
    ("fid_mean_per_iter", "fid_mean"),
    ("fid_std_per_iter", "fid_std"),
    ("L_loc_per_iter", "L_loc"),
    ("L_fid_per_iter", "L_fid"),
    ("neg_rtg_frac_per_iter", "neg_rtg_frac"),
    ("action_reward_corr_per_iter", "action_reward_corr"),
    ("rtg_frac_between_traj_var_per_iter", "rtg_frac_between_traj_var"),
    ("rtg_gpbelief_corr_per_iter", "rtg_gpbelief_corr"),
    ("grad_coherency_per_iter", "grad_coherency"),
    ("query_dist_to_xstar_per_iter", "query_dist_to_xstar"),
]
COLUMN_ORDER = [
    "method", "seed", "iteration", "regret", "cost", "cumulative_cost",
    "fidelity_H1_L0", "y_t", "x_t", "rtg_target", "btg_target",
    "fid_mean", "fid_std", "L_loc", "L_fid", "neg_rtg_frac",
    "action_reward_corr", "rtg_frac_between_traj_var", "rtg_gpbelief_corr",
    "grad_coherency", "query_dist_to_xstar",
]


def load_run(method, bm, seed):
    path = os.path.join(CKPT_DIR, f"{method}__{bm}__seed{seed}.json")
    if not os.path.exists(path):
        return None
    return json.load(open(path))


def rows_for_run(method, seed, d):
    present = {}
    n_iters = None
    for src_key, col in ARRAY_COLS:
        if src_key in d and d[src_key]:
            present[col] = d[src_key]
            n_iters = len(d[src_key]) if n_iters is None else n_iters
    if n_iters is None:
        return []
    rows = []
    for t in range(n_iters):
        row = {"method": method, "seed": seed, "iteration": t}
        for col in COLUMN_ORDER[3:]:
            vals = present.get(col)
            if vals is None or t >= len(vals):
                row[col] = None
            else:
                v = vals[t]
                row[col] = ",".join(f"{x:.4f}" for x in v) if isinstance(v, list) else v
        rows.append(row)
    return rows


def summary_row(method, bm, seed, d):
    rc = d.get("hf_regret_curve") or d.get("regret_curve") or []
    n_improved = sum(1 for i in range(1, len(rc)) if rc[i] < rc[i - 1] - 1e-12)
    distinct = len(set(f"{r:.6f}" for r in rc))
    return dict(
        benchmark=bm, method=method, seed=seed,
        n_iters=len(rc), final_regret=(rc[-1] if rc else None),
        incumbent_improved_count=n_improved,
        distinct_regret_values=distinct,
        lf_fraction=d.get("lf_fraction"),
        final_cost=(d["cost_curve"][-1] if d.get("cost_curve") else None),
    )


with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
    summary_rows = []
    for bm in BENCHMARKS:
        sheet_rows = []
        for method in METHODS:
            for seed in SEEDS:
                d = load_run(method, bm, seed)
                if d is None:
                    continue
                sheet_rows.extend(rows_for_run(method, seed, d))
                summary_rows.append(summary_row(method, bm, seed, d))
        if not sheet_rows:
            continue
        df = pd.DataFrame(sheet_rows, columns=COLUMN_ORDER)
        df.to_excel(writer, sheet_name=bm, index=False)

    pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Run Summary", index=False)

    legend = pd.DataFrame([
        ("method", "MF-DRO / SF-DRO / Greedy-MES"),
        ("seed", "random seed (42-46)"),
        ("iteration", "real BO iteration index (0-based; SF-DRO/Greedy-MES rows only have columns they actually save -- see below)"),
        ("regret", "simple regret at this iteration (running best; hf_regret_curve for MF-DRO, regret_curve for SF-DRO/Greedy-MES)"),
        ("cost", "post-init cumulative cost at this iteration"),
        ("cumulative_cost", "total cost including initialization (MF-DRO only)"),
        ("fidelity_H1_L0", "1=HF query, 0=LF query this iteration (MF-DRO, Greedy-MES only; SF-DRO is HF-only, not tracked)"),
        ("y_t", "raw observed value at this iteration's query (MF-DRO, Greedy-MES only)"),
        ("x_t", "query location this iteration, comma-joined per-dimension (MF-DRO, Greedy-MES only)"),
        ("rtg_target", "RTG target the DT was conditioned on (MF-DRO only)"),
        ("btg_target", "BTG target the DT was conditioned on (MF-DRO only)"),
        ("fid_mean", "mean P(HF) predicted by the fidelity head over this iteration's training batch (MF-DRO only)"),
        ("fid_std", "std of P(HF) over the training batch (MF-DRO only)"),
        ("L_loc", "location-head training loss this iteration (MF-DRO only)"),
        ("L_fid", "fidelity-head training loss this iteration (MF-DRO only)"),
        ("neg_rtg_frac", "fraction of rollout steps with negative RTG this iteration (MF-DRO only)"),
        ("action_reward_corr", "corr(RTG_tau, y_tau-incumbent) over the rollout batch (MF-DRO only; trajectory-level-luck diagnostic)"),
        ("rtg_frac_between_traj_var", "fraction of RTG variance explained by which trajectory a step came from vs. within-trajectory (MF-DRO only)"),
        ("rtg_gpbelief_corr", "corr(RTG_tau, GP posterior mean at x_tau) (MF-DRO only)"),
        ("grad_coherency", "mean pairwise cosine similarity of location-head gradients across an 8-rollout subsample (MF-DRO only)"),
        ("query_dist_to_xstar", "L2 distance from this iteration's real query to the benchmark's known optimum, normalized [0,1]^d (MF-DRO only; only populated for Hartmann_6D/Ackley_10D, which have a registered known_optimal_x)"),
        ("", ""),
        ("Run Summary sheet", "one row per (benchmark, method, seed): final regret, incumbent-improvement count, distinct regret values (freeze indicator), lf_fraction, final cost"),
        ("Note", "SF-DRO's saved result format only includes regret_curve/cost_curve -- all MF-DRO-specific diagnostic columns are blank for SF-DRO rows, not missing data"),
    ], columns=["Column", "Description"])
    legend.to_excel(writer, sheet_name="Legend", index=False)

# ---- Formatting pass ----
from openpyxl import load_workbook
wb = load_workbook(OUT_PATH)
header_font = Font(name="Arial", bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="4472C4")
body_font = Font(name="Arial", size=10)

for ws in wb.worksheets:
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
    ws.freeze_panes = "A2"
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max((len(str(c.value)) if c.value is not None else 0)
                      for c in ws[col_letter][:200])
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)

wb.save(OUT_PATH)
print(f"Saved {OUT_PATH}")
print(f"Sheets: {wb.sheetnames}")
